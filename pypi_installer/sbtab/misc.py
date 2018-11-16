#!/usr/bin/python
import re
import string
import libsbml
import numpy
import scipy
import scipy.linalg
import scipy.optimize
import random
import copy
import math
import os

try: from . import SBtab
except: import SBtab

def count_tabs(sbtab_string):
    '''
    count how many SBtabs are in in this string
    '''
    counter = 0
    for row in sbtab_string.split('\n'):
        if row.startswith('!!SBtab'):
            counter += 1
    return counter


def validate_file_extension(file_name, file_type):
    '''
    returns Boolean to evaluate if the file has the correct extension:
    sbml => xml
    sbtab => tsv, csv, xlsx
    '''
    # check extension for sbml file
    if file_type == 'sbml' and file_name[-3:] == 'xml': return True
    elif file_type == 'sbml': return False
    else: pass

    # check extension for sbtab file
    if file_type == 'sbtab' and file_name[-3:] == 'tsv': return True
    elif file_type == 'sbtab' and file_name[-3:] == 'csv': return True
    elif file_type == 'sbtab' and file_name[-4:] == 'xlsx': return True
    elif file_type == 'sbtab': return False
    else: pass

    # if something is completely off, return False
    return False


def check_delimiter(sbtab_file):
    '''
    determine the delimiter of the tabular file
    '''
    sep = False

    try:
        for row in sbtab_file.split('\n'):
            if row.startswith('!!'): continue
            if row.startswith('!'):
                s = re.search('(.)(!)', row[1:])
                # if there is only 1 column, we have to define a default separator
                # let's use a tab.
                try: sep = s.group(1)
                except: sep = '\t'
    except: pass

    return sep


def split_sbtabs(sbtab_strings):
    '''
    cuts an large SBtab string in single SBtab strings if necessary
    '''
    sbtabs = []
    sbtab_string = ''
    counter = 1
    
    for row in sbtab_strings.split('\n'):
        if row.startswith('!!'):
            if sbtab_string == '':
                sbtab_string = row + '\n'
                continue
            else:
                try:
                    sbtabs.append(sbtab_string)
                    sbtab_string = row + '\n'
                    counter += 1
                except:
                    print('Warning: Could not write SBtab %s' % counter)
                    counter += 1
        else:
            sbtab_string += row + '\n'
    sbtabs.append(sbtab_string) 
                    
    return sbtabs


def sbtab_to_html(sbtab, filename=None, mode='sbtab_online'):
    '''
    generates html view out of sbtab object
    'mode' can be 'sbtab_online' for generating HTML for the SBtab homepage or
    'standalone' for generating HTML as a standalone page
    '''
    def _is_float(value):
        '''
        checks if an element is a float in string format
        '''
        try:
            float(value)
            return True
        except: return False
        
    def _build_main(sbtab, sbtab_def):
        '''
        builds main body of HTML, which needs to be repeated
        for SBtab Documents
        '''
        no_link = ['(',')','+','-','<=>','or','and','FbcOr','FbcAnd']

        # get column descriptions for this table type and possible shortname links
        try: (col2description,col2link) = find_descriptions(sbtab_def, sbtab.table_type)
        except:
            col2description = False
            col2link = False
        
        # start main
        html = '<table class="table-striped">'

        # header row
        html += '<tr><th colspan="%s">%s</th></tr>' % (len(sbtab.columns), sbtab.header_row)

        # columns
        html += '<tr style="line-height:2;">'
        for col in sbtab.columns:
            try: title = col2description[col[1:]]
            except: title = ''
            html += '<th title="%s">%s</th>' % (title, col)
        html += '</tr>'

        # value rows
        for row in sbtab.value_rows:
            # set anchor for internal jump links
            try: html += '<tr id="%s" style="line-height:1.5;">' % row[sbtab.columns_dict['!ID']]
            except: html += '<tr style="line-height:1.5;">'
            for i,col in enumerate(row):
                # try and set internal jump links via shortnames
                try:
                    col2link[sbtab.columns[i]]
                    if col2link[sbtab.columns[i]] == 'True' and col != '' and col != False:
                        try:
                            html += '<td>'
                            split_column = col.split(' ')
                            for element in split_column:
                                if element not in no_link and not _is_float(element):
                                    html += '<a href="#%s">%s</a> ' % (element, element)
                                else:
                                    html += element + ' '
                            html += '</td>'
                        except: html += '<td>%s</td>' % (col)
                    else:
                        html += '<td>%s</td>' % (col)
                except:
                    if '!Identifiers' in sbtab.columns[i]:
                        try:
                            db = re.search('Identifiers:(.*)',sbtab.columns[i])
                            url = 'http://identifiers.org/%s/%s' % (db.group(1), col)                        
                            html += '<td><a href="%s">%s</a></td>' % (url, col)
                        except: html += '<td>%s</td>' % (col)
                    else:
                        html += '<td>%s</td>' % (col)
            html += '</tr>'

        # comment rows
        for row in sbtab.comments:
            html += '<tr>'
            for col in row:
                html += '<td>%s</td>' % col
            html += '</tr>'

        # close table
        html += '</table>'

        return html
    
    ##############################################################################
    # read in header and footer from HTML template
    if mode == 'sbtab_online':
        html_template = open('template_sbtab_online.html', 'r').read()
    elif mode == 'standalone':
        html_template = open('template_standalone.html', 'r').read()
    else:
        print('Invalid mode %s. Please use either "sbtab_online" or "standalone".' % mode)
        
    try:
        header = re.search('(<html lang="en">.*<main>)', html_template, re.DOTALL).group(0)
        footer = re.search('(</main>.*</html>)', html_template, re.DOTALL).group(0)
    except:
        print('Cannot read required template.html.')

    html = header

    # replace title placeholder with actual title
    html = html.replace('TitlePlaceholder',sbtab.filename)

    # read in definitions file for nice mouse over
    sbtab_def = open_definitions_file()
   
    # now build the html file
    if sbtab.object_type == 'table':
        html += _build_main(sbtab, sbtab_def)
    elif sbtab.object_type == 'doc':
        for sbtab in sbtab.sbtabs:
            html += _build_main(sbtab, sbtab_def) + '<br><hr>'
       
    else:
        print('The given SBtab object is invalid.')
            
    html += footer
    
    return html


def open_definitions_file(_path=None):
    '''
    open the SBtab definitions file, which can be in several locations
    '''
    sbtab_def = False
    
    if _path: try_paths = [_path]
    else:
        try_paths = ['definitions.tsv',
                     os.path.join(os.path.dirname(__file__), '../static/files/default_files/definitions.tsv'),
                     os.path.join(os.path.dirname(__file__), '../definition_table/definitions.tsv'),
                     os.path.join(os.path.dirname(__file__), 'definitions.tsv')]
    
    for path in try_paths:
        try:
            def_file = open(path, 'r')
            sbtab_def = SBtab.SBtabTable(def_file.read(), 'definitions.tsv')
            def_file.close()
            break
        except: pass

    return sbtab_def


def find_descriptions(def_file, table_type):
    '''
    preprocesses the definition file in order to enable some nice mouseover effects for the known column names
    '''
    col2description = {}
    col2link = {}

    for row in def_file.value_rows:
        if row[def_file.columns_dict['!IsPartOf']] == table_type:
            col2description[row[def_file.columns_dict['!ComponentName']]] = row[def_file.columns_dict['!Description']]
            col2link['!'+row[def_file.columns_dict['!ComponentName']]] = row[def_file.columns_dict['!linksShortname']]

    return (col2description,col2link)


def xlsx_to_tsv(file_object, f='web'):
    '''
    convert xlsx SBtab file to tsv format
    '''
    import openpyxl
    from io import BytesIO

    if f == 'web': wb = openpyxl.load_workbook(filename = BytesIO(file_object))
    else: wb = openpyxl.load_workbook(filename = file_object)
    ws = wb.active
    ranges = wb[ws.title]
    table_string = ''

    for row in ranges:
        for column in row:
            if str(row[0].value).startswith('!'):
                if column.value != None and str(column.value) != '':
                    table_string += str(column.value) + '\t'
            else:
                table_string += str(column.value) + '\t'
        table_string = table_string[:-1] + '\n'
        
    return table_string


def tab_to_xlsx(sbtab_object):
    '''
    converts SBtab object to xlsx file object
    '''
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    
    ws['A1'] = sbtab_object.header_row
    ws.append(sbtab_object.columns)
    for row in sbtab_object.value_rows:
        ws.append(row)

    wb.save('transition.xlsx')
    
    fileobject = open('transition.xlsx','rb')

    return fileobject.read()


def xml_to_html(sbml_file):
    '''
    generates html view out of xml file
    '''
    old_sbml = sbml_file.split('\n')
    new_sbml = '<xmp>'
    for row in old_sbml:
        new_sbml += row + '\n'
    new_sbml += '</xmp>'

    return new_sbml

            
def extract_supported_table_types():
    '''
    extracts all allowed SBtab TableTypes from the definition file
    '''
    sbtab_def = open_definitions_file()
    
    supported_types = []
    for row in sbtab_def.value_rows:
        t = row[sbtab_def.columns_dict['!IsPartOf']]
        if t not in supported_types:
            supported_types.append(t)

    return supported_types
