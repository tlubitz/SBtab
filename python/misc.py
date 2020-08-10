#!/usr/bin/python
'''
A collection of useful tools and functions for the manipulation of SBtab tables and SBtab documents.
'''
import re
import string
import libsbml
import numpy
import scipy
import scipy.linalg
import scipy.optimize
import random
import copy
import math
import os

try: from . import SBtab
except: import SBtab

def count_tabs(sbtab_string):
    '''
    Counts how many SBtabs are in in a given string.

    Parameters
    ----------
    sbtab_string: str
        SBtab table or tables in string representation.
        
    Returns: int
        Amount of SBtab tables in given string.
    '''
    counter = 0
    for row in sbtab_string.split('\n'):
        if row.startswith('!!SBtab') or row.startswith('!!ObjTables'):
            counter += 1
    return counter


def validate_file_extension(file_name, file_type):
    '''
    Returns Boolean flag to evaluate if the file has the correct extension:
    sbml => xml
    sbtab => tsv, csv, xlsx.

    Parameters
    ----------
    file_name: str
        Name of the file.
    file_type: str
        Type of the file ('sbtab' or 'sbml').

    Returns: Bool
        Boolean flag indicating if the given file extension corresponds to the given file type.
    '''
    # check extension for sbml file
    if file_type == 'sbml' and file_name[-3:] == 'xml': return True
    elif file_type == 'sbml': return False
    else: pass

    # check extension for sbtab file
    if file_type == 'sbtab' and file_name[-3:] == 'tsv': return True
    elif file_type == 'sbtab' and file_name[-3:] == 'csv': return True
    elif file_type == 'sbtab' and file_name[-4:] == 'xlsx': return True
    elif file_type == 'sbtab': return False
    else: pass

    # if something is completely off, return False
    return False


def check_delimiter(sbtab_file):
    '''
    Determines the delimiter of the SBtab table

    Parameters
    ----------
    sbtab_file: str
        SBtab table in string representation.

    Returns: str
        Delimiter of the SBtab table ('\t', ',', or ';')
    '''
    sep = False

    try:
        for row in sbtab_file.split('\n'):
            if row.startswith('!!'): continue
            if row.startswith('!'):
                s = re.search('(.)(!)', row[1:])
                # if there is only 1 column, we have to define a default separator
                # let's use a tab.
                try: sep = s.group(1)
                except: sep = '\t'
    except: pass

    return sep


def split_sbtabs(sbtab_strings):
    '''
    Cuts one SBtab string in single SBtab strings if necessary.

    Parameters
    ----------
    sbtab_strings: str
        SBtab table or tables in string representation.

    Returns: list
        List of SBtab tables in string representation.
    '''
    sbtabs = []
    sbtab_string = ''
    counter = 1
    
    for row in sbtab_strings.split('\n'):
        if row.startswith('!!!') or row.startswith('"!!!'): continue
        if row.startswith('!!'):
            if sbtab_string == '':
                sbtab_string = row + '\n'
                continue
            else:
                try:
                    if sbtab_string.startswith('!!SBtab') or sbtab_string.startswith('!!ObjTables'):
                        sbtabs.append(sbtab_string)
                        counter += 1
                    sbtab_string = row + '\n'
                except:
                    print('Warning: Could not write SBtab %s' % counter)
                    counter += 1
        else:
            sbtab_string += row + '\n'

    if sbtab_string.startswith('!!SBtab') or sbtab_string.startswith('!!ObjTables'):
        sbtabs.append(sbtab_string)
                    
    return sbtabs


def sbtab_to_html(sbtab, filename=None, mode='sbtab_online', template = [], put_links = True, title_string='', show_header_row = True, show_table_name = False, show_table_text = False, definitions_file=''):
    '''
    Generates html view out of SBtab table or SBtab document object.

    Parameters
    ----------
    sbtab: SBtab.SBtabTable | SBtab.SBtabDocument
        Either SBtab table object or SBtab document object.
    filename: str
        File name of the SBtab table.
    mode: str
        Defines the type of HTML to be generated ('sbtab_online' for the SBtab online
        interface or 'standalone' for a sole HTML page without online binding).

    Returns: str
        SBtab object as HTML string.
    '''
    def _is_float(value):
        '''
        checks if an element is a float in string format
        '''
        try:
            float(value)
            return True
        except: return False
        
    def _build_main(sbtab, sbtab_def):
        '''
        builds main body of HTML, which needs to be repeated
        for SBtab Documents
        '''
        no_link = ['(',')','+','-','<=>','or','and','FbcOr','FbcAnd']
        # get column descriptions for this table type and possible shortname links
        try: (col2description,col2link) = find_descriptions(sbtab_def, sbtab.table_type)
        except:
            col2description = False
            col2link = False
        
        # start main
        html = '<table class="table-striped">'

        # table name
        if show_table_name:
            html += '<center><h2>%s</h2></center>' % (sbtab.get_attribute('TableName'))

        if show_table_text:
            if len(sbtab.get_attribute('Text')):
                html += '<center><p>%s</p></center>' % (sbtab.get_attribute('Text'))

        # header row
        #html += '<thead><tr><th colspan="%s">%s</th></tr></thead>' % (len(sbtab.columns), sbtab.header_row)
        if show_header_row:
            html += '<h4>%s</h4>' % (sbtab.header_row)

        # columns
        html += '<thead>'
        html += '<tr style="line-height:2;">'
        for col in sbtab.columns:
            try: title = col2description[col[1:]]
            except: title = ''
            html += '<th title="%s">%s</th>' % (title, col)
        html += '</tr>'
        html += '</thead>'
        html += '<tbody>'

        # value rows
        for row in sbtab.value_rows:
            # set anchor for internal jump links
            try: html += '<tr id="%s" style="line-height:1.5;">' % row[sbtab.columns_dict['!ID']]
            except: html += '<tr style="line-height:1.5;">'
            for i,col in enumerate(row):
                # try and set internal jump links via shortnames
                try:
                    col2link[sbtab.columns[i]]
                    if col2link[sbtab.columns[i]] == 'True' and col != '' and col != False:
                        try:
                            html += '<td>'
                            split_column = col.split(' ')
                            for element in split_column:
                                if element not in no_link and not _is_float(element) and put_links:
                                    #html += '<a href="#%s">%s</a> ' % (element, element)    #internal links
                                    html += element
                                else:
                                    html += element + ' '
                            html += '</td>'
                        except: html += '<td>%s</td>' % (col)
                    else:
                        html += '<td>%s</td>' % (col)
                except:
                    if '!Identifiers' in sbtab.columns[i]:
                        try:
                            db = re.search('Identifiers:(.*)',sbtab.columns[i])
                            url = 'http://identifiers.org/%s/%s' % (db.group(1), col)                        
                            html += '<td><a href="%s">%s</a></td>' % (url, col)
                        except: html += '<td>%s</td>' % (col)
                    else:
                        html += '<td>%s</td>' % (col)
            html += '</tr>'

        # comment rows
        for row in sbtab.comments:
            html += '<tr>'
            for col in row:
                html += '<td>%s</td>' % col
            html += '</tr>'

        # close table
        html += '</tbody></table>'

        return html
    
    ##############################################################################
    # read in header and footer from HTML template
    if mode == 'sbtab_online':
        p = os.path.join(os.path.dirname(__file__), '../modules/template_sbtab_online.html')
        try:
            html = open(p, 'r')
            html_template = html.read()
            html.close()
        except:
            print('HTML template was not found.')
            return False
    elif mode == 'standalone':
        html_template = False
        try_paths = ['html_templates/template_standalone.html',                     
                     os.path.join(os.path.dirname(__file__), '../html_templates/template_standalone.html'),
                     os.path.join(os.path.dirname(__file__), 'html_templates/template_standalone.html'),
                     template]
        for path in try_paths:
            try:
                html = open(path, 'r')
                html_template = html.read()
                html.close()
            except: pass
        if not html_template:
            print('HTML template was not found.')
            return False
    else:
        print('Invalid mode %s. Please use either "sbtab_online" or "standalone".' % mode)
        return False

    try:
        header = re.search('(<html lang="en">.*<main>)', html_template, re.DOTALL).group(0)
        footer = re.search('(</main>.*</html>)', html_template, re.DOTALL).group(0)
    except:
        print('Cannot read required template.html.')
        return False

    html = header

    try:
        ot = sbtab.object_type
    except:
        print('You have not provided a valid SBtab object as input.')
        return False
    
    # replace title placeholder with actual title
    if len(title_string):
        html = html.replace('TitlePlaceholder',title_string)
    else:
        html = html.replace('TitlePlaceholder',sbtab.filename)

    # read in definitions file for nice mouse over
    if mode == 'standalone' and len(definitions_file):
        sbtab_def = open_definitions_file(definitions_file)
    else:
        sbtab_def = open_definitions_file()
        
    # now build the html file
    if sbtab.object_type == 'table':
        html += _build_main(sbtab, sbtab_def)
    elif sbtab.object_type == 'doc':
        for sbtab in sbtab.sbtabs:
            html += _build_main(sbtab, sbtab_def) + '<br><hr>'
    else:
        print('The given SBtab object is invalid.')
        return False
            
    html += footer
    
    return html


def open_definitions_file(_path=None):
    '''
    Opens the SBtab definitions file, which can be in several locations.

    Parameters
    ----------
    _path: str
        Optional path to the definitions.tsv.

    Returns: SBtab.SBtabTable
        SBtab definitions file as SBtabTable object.        
    '''
    sbtab_def = False
    
    if _path: try_paths = [_path]
    else:
        try_paths = ['definitions.tsv',
                     os.path.join(os.path.dirname(__file__), '../static/files/default_files/definitions.tsv'),
                     os.path.join(os.path.dirname(__file__), '../definition_table/definitions.tsv'),
                     os.path.join(os.path.dirname(__file__), 'definitions.tsv')]
    
    for path in try_paths:
        try:
            def_file = open(path, 'r')
            file_content = def_file.read()
            sbtab_def = SBtab.SBtabTable(file_content, 'definitions.tsv')
            def_file.close()
            break
        except: pass

    return sbtab_def

            
def check_obj(file_string):
    '''
    Tests a file string if it is SBtab or ObjTables format.

    Parameters
    ----------
    file_string: str
        Content of a read file.

    Returns: Boolean
        True if ObjTables
        False if SBtab
    '''
    objTables = False
    for row in file_string:
        if row.startswith('!!!ObjTables') or row.startswith('!!ObjTables'):
            objTables = True
    return objTables


def extract_supported_table_types():
    '''
    Extracts all allowed SBtab table types from the definitions file.

    Returns: list
        List of supported SBtab table types.
    '''
    sbtab_def = open_definitions_file()
    
    supported_types = []
    for row in sbtab_def.value_rows:
        t = row[sbtab_def.columns_dict['!Parent']]
        if t not in supported_types and t != 'SBtab':
            supported_types.append(t)

    return supported_types


def find_descriptions(def_file, table_type):
    '''
    Preprocesses the definitions file in order to enable some nice mouseover effects for the known column names.

    Parameters
    ----------
    def_file: SBtab.SBtabTable
        Definitions file as SBtab table object.
    table_type: str
        SBtab table type for which the descriptions shall be extracted.

    Returns: (dict, dict)
        Two dictionaries that link the column name to its description and to a Bool flag indicating if it
        is a shortname identifier possibly linking to another SBtab table.
    '''
    col2description = {}
    col2link = {}

    for row in def_file.value_rows:
        if row[def_file.columns_dict['!Parent']] == table_type:
            col2description[row[def_file.columns_dict['!Name']]] = row[def_file.columns_dict['!Description']]
            col2link['!'+row[def_file.columns_dict['!Name']]] = row[def_file.columns_dict['!isShortname']]

    return (col2description, col2link)


def xlsx_to_tsv(file_object, f='web'):
    '''
    Converts xlsx SBtab file to tsv format.

    Parameters
    ----------
    file_object: xlsx file object
        SBtab table as xlsx file object.
    f: str
        String indicating how the file object is represented ('web' for online file handling, otherwise normal file handling)
    
    Returns: str
        SBtab table as tsv string.
    '''
    import openpyxl
    from io import BytesIO

    if f == 'web': wb = openpyxl.load_workbook(filename = BytesIO(file_object))
    else: wb = openpyxl.load_workbook(filename = file_object)
    ws = wb.active
    ranges = wb[ws.title]
    table_string = ''

    for row in ranges:
        for column in row:
            if str(row[0].value).startswith('!'):
                if column.value != None and str(column.value) != '':
                    table_string += str(column.value) + '\t'
            else:
                table_string += str(column.value) + '\t'
        table_string = table_string[:-1] + '\n'
        
    return table_string


def tab_to_xlsx(sbtab_object):
    '''
    Converts SBtab object to xlsx file object.

    Parameters
    ----------
    sbtab_object: SBtab.SBtabTable
        SBtab table object.

    Returns: xlsx file object
        SBtab table as xlsx file object.
    '''
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    
    ws['A1'] = sbtab_object.header_row
    ws.append(sbtab_object.columns)
    for row in sbtab_object.value_rows:
        ws.append(row)

    wb.save('transition.xlsx')
    
    f = open('transition.xlsx','rb')
    fileobject = f.read()
    f.close()

    return fileobject


def xml_to_html(sbml_file):
    '''
    Generates HTML view of XML (SBML) file.

    Parameters
    ----------
    sbml_file: str
        SBML file in string representation.

    Returns: str
        SBML file as HTML view.
    '''
    old_sbml = sbml_file.split('\n')
    new_sbml = '<xmp>'
    for row in old_sbml:
        new_sbml += row + '\n'
    new_sbml += '</xmp>'

    return new_sbml

